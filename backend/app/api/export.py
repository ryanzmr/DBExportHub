import os
import pandas as pd
import tempfile
import uuid
import pathlib
from datetime import datetime
from typing import List, Dict, Any, Optional

from ..config import settings
from ..database import get_db_connection, execute_query, query_to_dataframe


def preview_data(params):
    """
    Generate a preview of the data based on the export parameters.
    Returns a limited number of records for preview in the UI.
    """
    try:
        # Connect to the database
        with get_db_connection(
            params.server, params.database, params.username, params.password
        ) as conn:
            # Build the query parameters for the stored procedure
            sp_params = {
                "fromMonth": params.fromMonth,
                "ToMonth": params.toMonth,
                "hs": params.hs,
                "prod": params.prod,
                "Iec": params.iec,
                "ExpCmp": params.expCmp,
                "forcount": params.forcount,
                "forname": params.forname,
                "port": params.port
            }
            
            # Execute the stored procedure
            # First, execute the stored procedure to populate the temp table
            param_list = list(sp_params.values())
            sp_call = f"EXEC {settings.EXPORT_STORED_PROCEDURE} ?, ?, ?, ?, ?, ?, ?, ?, ?"
            execute_query(conn, sp_call, param_list)
            
            # Then query the temp table for preview data
            preview_query = f"SELECT TOP {params.max_records} * FROM {settings.EXPORT_TEMP_TABLE}"
            result = execute_query(conn, preview_query)
            
            return result
    except Exception as e:
        raise Exception(f"Error generating preview: {str(e)}")


def generate_excel(params):
    """
    Generate an Excel file based on the export parameters.
    Returns the path to the generated Excel file.
    """
    try:
        # Connect to the database
        with get_db_connection(
            params.server, params.database, params.username, params.password
        ) as conn:
            # Build the query parameters for the stored procedure
            sp_params = {
                "fromMonth": params.fromMonth,
                "ToMonth": params.toMonth,
                "hs": params.hs,
                "prod": params.prod,
                "Iec": params.iec,
                "ExpCmp": params.expCmp,
                "forcount": params.forcount,
                "forname": params.forname,
                "port": params.port
            }
            
            # Execute the stored procedure
            # First, execute the stored procedure to populate the temp table
            param_list = list(sp_params.values())
            sp_call = f"EXEC {settings.EXPORT_STORED_PROCEDURE} ?, ?, ?, ?, ?, ?, ?, ?, ?"
            execute_query(conn, sp_call, param_list)
            
            # Then query the temp table for all data
            df = query_to_dataframe(conn, f"SELECT * FROM {settings.EXPORT_TEMP_TABLE}")
            
            # Ensure temp directory exists - explicitly create it here to avoid issues
            temp_dir = pathlib.Path(settings.TEMP_DIR)
            os.makedirs(temp_dir, exist_ok=True)
            
            # Generate a unique filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            unique_id = str(uuid.uuid4())[:8]
            filename = f"export_{timestamp}_{unique_id}.xlsx"
            file_path = str(temp_dir / filename)
            
            # Ensure the file path is an absolute path
            file_path = os.path.abspath(file_path)
            
            # Check if template exists
            template_path = pathlib.Path(settings.EXCEL_TEMPLATE_PATH)
            if template_path.exists():
                # Use the template - properly copy the template file first
                import shutil
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter
                
                try:
                    # Copy the template file to the destination
                    shutil.copy2(template_path, file_path)
                    
                    # Load the workbook from the copied template
                    workbook = openpyxl.load_workbook(file_path)
                    
                    # Select the first worksheet or create a new one
                    if len(workbook.sheetnames) > 0:
                        worksheet = workbook.active
                    else:
                        worksheet = workbook.create_sheet(title="Export Data")
                    
                    # Define header style
                    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    thin_border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
                    
                    # Write headers
                    for col_idx, column_name in enumerate(df.columns, 1):
                        cell = worksheet.cell(row=1, column=col_idx, value=column_name)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    
                    # Write data
                    for row_idx, row in enumerate(df.values, 2):
                        for col_idx, value in enumerate(row, 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                            cell.border = thin_border
                            
                            # Format date columns
                            if col_idx == 2 or col_idx == 46:  # SB_Date and LEO_Date columns
                                if value is not None:
                                    cell.number_format = 'dd-mm-yyyy'
                            
                            # Format numeric columns
                            if isinstance(value, (int, float)) and col_idx not in [2, 46]:  # Not date columns
                                if 'Value' in df.columns[col_idx-1] or 'Rate' in df.columns[col_idx-1]:
                                    cell.number_format = '#,##0.00'
                                else:
                                    cell.number_format = '#,##0'
                    
                    # Auto-adjust column widths
                    for col_idx, column in enumerate(df.columns, 1):
                        column_width = max(len(str(column)), df[column].astype(str).str.len().max())
                        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(column_width + 2, 40)
                    
                    # Freeze the header row
                    worksheet.freeze_panes = 'A2'
                    
                    # Save the workbook
                    workbook.save(file_path)
                    
                except Exception as e:
                    raise Exception(f"Error using template: {str(e)}")
            else:
                # Create a new Excel file without template
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Export Data', index=False)
                    
                    # Get the workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Export Data']
                    
                    # Auto-adjust column widths
                    for col_idx, column in enumerate(df.columns, 1):
                        column_width = max(len(str(column)), df[column].astype(str).str.len().max())
                        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(column_width + 2, 40)
                    
                    # Freeze the header row
                    worksheet.freeze_panes = 'A2'
            
            return file_path
    except Exception as e:
        raise Exception(f"Error generating Excel file: {str(e)}")